import os
import unittest
from openpyxl import load_workbook
from algorithmic_paper_trader import load_env, run_polling_mvp, run_simulation


class TestPaperTrader(unittest.TestCase):
    def test_env_loader(self):
        env = load_env(os.path.join(os.getcwd(), ".env"))
        self.assertIn("MODE", env)
        self.assertIn("POLL_INTERVAL_SECONDS", env)
        self.assertEqual(os.environ.get("MODE"), env["MODE"])

    def test_polling_mvp_alerts_and_excel(self):
        os.environ["MODE"] = "mock"
        alerts, path = run_polling_mvp(iterations=3, poll_interval_seconds=0)
        self.assertTrue(len(alerts) >= 2)
        self.assertTrue(any(a.startswith("ALERT:") for a in alerts))
        self.assertTrue(any("₹" in a for a in alerts))
        self.assertTrue(os.path.exists(path))
        wb = load_workbook(path)
        ws = wb["Transactions"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers, [
            "Order Date",
            "Order Time",
            "Order Executed Time",
            "Symbol",
            "Instrument",
            "Transaction",
            "Quantity",
            "Price",
            "Order ID",
            "Order Status",
            "Strategy",
            "PnL",
            "Remarks",
        ])
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertTrue(len(rows) >= 2)
        # Check formatting for Price and PnL columns
        price_cell = ws["H2"]
        pnl_cell = ws["L2"]
        self.assertIn("₹", price_cell.number_format)
        self.assertIn("₹", pnl_cell.number_format)
        # Summary sheet exists
        self.assertIn("Summary", wb.sheetnames)
        summary = wb["Summary"]
        self.assertEqual(summary["A1"].value, "Metric")
        self.assertEqual(summary["B1"].value, "Value")
        self.assertEqual(summary["A3"].value, "Total PnL")


if __name__ == "__main__":
    unittest.main()
