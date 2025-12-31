import os
import uuid
from dataclasses import dataclass, field
from datetime import datetime, time
from typing import List, Dict, Optional, Tuple
import math
import pandas as pd
from openpyxl import Workbook
import time as time_module
import json
import urllib.request
import urllib.parse
import gzip
import io


@dataclass
class Config:
    capital: float
    allocation_pct: float
    start_time: time
    stop_entry_time: time
    square_off_time: time
    exchanges: List[str]
    strategy_name: str = "52W Breakout"
    min_price: float = 50.0
    mode: str = "mock"


@dataclass
class Instrument:
    symbol: str
    exchange: str


@dataclass
class Position:
    symbol: str
    exchange: str
    side: str
    qty: int
    entry_price: float
    entry_time: datetime
    open: bool = True


@dataclass
class OrderLogEntry:
    order_date: str
    order_time: str
    executed_time: str
    symbol: str
    instrument: str
    transaction: str
    quantity: int
    price: float
    order_id: str
    order_status: str
    strategy: str
    pnl: float
    remarks: str


class MarketDataProvider:
    def __init__(self, mode: str = "mock"):
        self.mode = mode
        self.mock_prices: Dict[Tuple[str, str], float] = {}
        self.mock_high_low: Dict[Tuple[str, str], Tuple[float, float]] = {}
        self.last_prices: Dict[Tuple[str, str], Optional[float]] = {}
        self._cache_52w: Dict[Tuple[str, str], Tuple[float, float]] = {}

    def set_mock(self, symbol: str, exchange: str, ltp: float, high_52w: float, low_52w: float):
        k = (symbol, exchange)
        self.mock_prices[k] = ltp
        self.mock_high_low[k] = (high_52w, low_52w)
        self.last_prices.setdefault(k, None)

    def update_mock_price(self, symbol: str, exchange: str, ltp: float):
        k = (symbol, exchange)
        self.last_prices[k] = self.mock_prices.get(k)
        self.mock_prices[k] = ltp

    def get_ltp(self, symbol: str, exchange: str) -> float:
        if self.mode == "mock":
            return self.mock_prices[(symbol, exchange)]
        if self.mode == "upstox":
            ik = self._instrument_key(symbol, exchange)
            if ik is None:
                raise RuntimeError(f"Missing instrument_key for {exchange}:{symbol}")
            base = os.environ.get("UPSTOX_BASE_URL", "https://api.upstox.com")
            url = f"{base}/v2/market-quote/ltp?instrument_key={urllib.parse.quote(ik)}"
            data = self._http_get_json(url)
            data_obj = data.get("data", {})
            key = next(iter(data_obj.keys()), None)
            if not key:
                raise RuntimeError("Invalid LTP response")
            return float(data_obj[key]["last_price"])
        raise NotImplementedError("Data access mode not configured")

    def get_last_price(self, symbol: str, exchange: str) -> Optional[float]:
        return self.last_prices.get((symbol, exchange))

    def get_52w_high_low(self, symbol: str, exchange: str) -> Tuple[float, float]:
        if self.mode == "mock":
            return self.mock_high_low[(symbol, exchange)]
        if self.mode == "upstox":
            k = (symbol, exchange)
            if k in self._cache_52w:
                return self._cache_52w[k]
            ik = self._instrument_key(symbol, exchange)
            if ik is None:
                raise RuntimeError(f"Missing instrument_key for {exchange}:{symbol}")
            base = os.environ.get("UPSTOX_BASE_URL", "https://api.upstox.com")
            to_date = datetime.today().date().strftime("%Y-%m-%d")
            from_date = (datetime.today().date() - pd.Timedelta(days=365)).strftime("%Y-%m-%d")
            url = f"{base}/v2/historical-candle/{urllib.parse.quote(ik)}/day/{to_date}/{from_date}"
            data = self._http_get_json(url)
            candles = data.get("data", {}).get("candles", [])
            if not candles:
                raise RuntimeError("No historical candles received")
            highs = [c[2] for c in candles]
            lows = [c[3] for c in candles]
            hi, lo = max(highs), min(lows)
            self._cache_52w[k] = (hi, lo)
            return hi, lo
        raise NotImplementedError("Data access mode not configured")

    def _instrument_key(self, symbol: str, exchange: str) -> Optional[str]:
        key = self._resolve_instrument_key_online(symbol, exchange)
        if key:
            return key
        return None

    def _resolve_instrument_key_online(self, symbol: str, exchange: str) -> Optional[str]:
        seg = "NSE_EQ" if exchange.upper() == "NSE" else "BSE_EQ"
        url = "https://upstox.com/developer/api-documentation/instruments/"
        try:
            doc = self._http_get_text(url)
        except Exception:
            return None
        candidates = []
        for href in self._extract_links(doc):
            if ("NSE" in href or "BSE" in href) and ("json" in href or "gz" in href):
                candidates.append(href)
        for link in candidates:
            try:
                data = self._download_instruments_json(link)
            except Exception:
                continue
            for item in data:
                if (
                    str(item.get("segment", "")).upper() == seg
                    and str(item.get("instrument_type", "")).upper() == "EQ"
                    and str(item.get("trading_symbol", "")).upper() == symbol.upper()
                ):
                    ik = item.get("instrument_key")
                    if ik:
                        return ik
        return None

    def _http_get_text(self, url: str) -> str:
        req = urllib.request.Request(url, headers={"Accept": "text/html"}, method="GET")
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.read().decode("utf-8", errors="ignore")

    def _extract_links(self, html: str) -> List[str]:
        links: List[str] = []
        for part in html.split('"'):
            if part.startswith("http") and ("json" in part or "gz" in part):
                links.append(part)
        for part in html.split("'"):
            if part.startswith("http") and ("json" in part or "gz" in part):
                links.append(part)
        return list(set(links))

    def _download_instruments_json(self, url: str) -> List[Dict]:
        req = urllib.request.Request(url, headers={"Accept": "application/json"}, method="GET")
        with urllib.request.urlopen(req, timeout=20) as resp:
            content = resp.read()
        if url.endswith(".gz"):
            with gzip.GzipFile(fileobj=io.BytesIO(content)) as gz:
                content = gz.read()
        text = content.decode("utf-8")
        data = json.loads(text)
        if isinstance(data, dict):
            for k in ("data", "instruments"):
                if k in data and isinstance(data[k], list):
                    return data[k]
            return []
        if isinstance(data, list):
            return data
        return []

    def _http_get_json(self, url: str) -> Dict:
        token = os.environ.get("UPSTOX_ACCESS_TOKEN", "")
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
        if token:
            headers["Authorization"] = f"Bearer {token}"
        req = urllib.request.Request(url, headers=headers, method="GET")
        with urllib.request.urlopen(req, timeout=10) as resp:
            body = resp.read()
            return json.loads(body.decode("utf-8"))


class RiskManager:
    def __init__(self, config: Config):
        self.config = config
        self.capital_remaining = config.capital

    def eligible(self, price: float) -> bool:
        return price >= self.config.min_price and self.capital_remaining >= self.config.capital * self.config.allocation_pct

    def compute_qty(self, price: float) -> int:
        alloc = self.config.capital * self.config.allocation_pct
        qty = math.floor(alloc / price)
        return max(qty, 0)

    def reserve_capital(self, price: float, qty: int):
        self.capital_remaining -= price * qty

    def release_capital(self, price: float, qty: int):
        self.capital_remaining += price * qty


class Strategy52WBreakout:
    def __init__(self, mdp: MarketDataProvider):
        self.mdp = mdp

    def signal(self, symbol: str, exchange: str) -> Optional[str]:
        ltp = self.mdp.get_ltp(symbol, exchange)
        prev = self.mdp.get_last_price(symbol, exchange)
        high_52w, low_52w = self.mdp.get_52w_high_low(symbol, exchange)
        if prev is not None and prev <= high_52w and ltp > high_52w:
            return "BUY"
        if prev is not None and prev >= low_52w and ltp < low_52w:
            return "SELL"
        return None


class PaperBroker:
    def __init__(self, config: Config, mdp: MarketDataProvider, risk: RiskManager):
        self.config = config
        self.mdp = mdp
        self.risk = risk
        self.positions: Dict[Tuple[str, str], Position] = {}
        self.logs: List[OrderLogEntry] = []
        self.events: List[str] = []

    def can_enter_now(self, now: datetime) -> bool:
        t = now.time()
        return self.config.start_time <= t <= self.config.stop_entry_time

    def must_square_off(self, now: datetime) -> bool:
        t = now.time()
        return t >= self.config.square_off_time

    def place_entry(self, symbol: str, exchange: str, side: str, now: datetime):
        ltp = self.mdp.get_ltp(symbol, exchange)
        if not self.risk.eligible(ltp):
            return
        qty = self.risk.compute_qty(ltp)
        if qty < 1:
            return
        self.risk.reserve_capital(ltp, qty)
        pos = Position(symbol=symbol, exchange=exchange, side=side, qty=qty, entry_price=ltp, entry_time=now)
        self.positions[(symbol, exchange)] = pos
        oid = str(uuid.uuid4())
        self.logs.append(OrderLogEntry(
            order_date=now.strftime("%Y-%m-%d"),
            order_time=now.strftime("%H:%M:%S"),
            executed_time=now.strftime("%H:%M:%S"),
            symbol=symbol,
            instrument=f"{exchange}:{symbol}",
            transaction=side,
            quantity=qty,
            price=ltp,
            order_id=oid,
            order_status="FILLED",
            strategy=self.config.strategy_name,
            pnl=0.0,
            remarks="Strategy trigger: cross 52W boundary"
        ))
        self.events.append(f"ALERT: [{now.strftime('%H:%M:%S')}] {exchange}:{symbol} {side} at ₹{ltp:.2f} qty {qty} — 52W breakout")

    def try_exit_targets(self, now: datetime):
        for k, pos in list(self.positions.items()):
            ltp = self.mdp.get_ltp(pos.symbol, pos.exchange)
            target = pos.entry_price * (1.01 if pos.side == "BUY" else 0.99)
            hit = ltp >= target if pos.side == "BUY" else ltp <= target
            if hit:
                self.execute_exit(pos, now, ltp, "Profit target hit")

    def square_off_all(self, now: datetime):
        for k, pos in list(self.positions.items()):
            ltp = self.mdp.get_ltp(pos.symbol, pos.exchange)
            self.execute_exit(pos, now, ltp, "Time-based square-off")

    def execute_exit(self, pos: Position, now: datetime, ltp: float, reason: str):
        exit_side = "SELL" if pos.side == "BUY" else "BUY"
        pnl = (ltp - pos.entry_price) * pos.qty if pos.side == "BUY" else (pos.entry_price - ltp) * pos.qty
        self.risk.release_capital(pos.entry_price, pos.qty)
        oid = str(uuid.uuid4())
        self.logs.append(OrderLogEntry(
            order_date=now.strftime("%Y-%m-%d"),
            order_time=now.strftime("%H:%M:%S"),
            executed_time=now.strftime("%H:%M:%S"),
            symbol=pos.symbol,
            instrument=f"{pos.exchange}:{pos.symbol}",
            transaction=exit_side,
            quantity=pos.qty,
            price=ltp,
            order_id=oid,
            order_status="FILLED",
            strategy=self.config.strategy_name,
            pnl=round(pnl, 2),
            remarks=reason
        ))
        pos.open = False
        del self.positions[(pos.symbol, pos.exchange)]
        self.events.append(f"ALERT: [{now.strftime('%H:%M:%S')}] {pos.exchange}:{pos.symbol} {exit_side} at ₹{ltp:.2f} qty {pos.qty} — {reason} — PnL ₹{round(pnl,2):.2f}")


class ExcelLogger:
    def __init__(self, path: str):
        self.path = path

    def write(self, entries: List[OrderLogEntry]):
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        headers = [
            "Order Date",
            "Order Time",
            "Order Executed Time",
            "Symbol",
            "Instrument",
            "Transaction",
            "Quantity",
            "Price",
            "Order ID",
            "Order Status",
            "Strategy",
            "PnL",
            "Remarks",
        ]
        ws.append(headers)
        for e in entries:
            ws.append([
                e.order_date,
                e.order_time,
                e.executed_time,
                e.symbol,
                e.instrument,
                e.transaction,
                e.quantity,
                e.price,
                e.order_id,
                e.order_status,
                e.strategy,
                e.pnl,
                e.remarks,
            ])
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        for col in ["A","B","C","D","E","F","G","H","I","J","K","L","M"]:
            ws.column_dimensions[col].width = 20
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
            for cell in row:
                cell.number_format = u'₹#,##0.00'
        for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
            for cell in row:
                cell.number_format = u'₹#,##0.00'
        summary = wb.create_sheet(title="Summary")
        total_trades = len([e for e in entries if e.transaction in ("SELL","BUY")])
        total_pnl = sum(e.pnl for e in entries if e.transaction in ("SELL","BUY"))
        wins = len([e for e in entries if e.remarks == "Profit target hit"])
        time_exits = len([e for e in entries if e.remarks == "Time-based square-off"])
        summary.append(["Metric", "Value"])
        summary.append(["Total Trades (entries+exits)", total_trades])
        summary.append(["Total PnL", total_pnl])
        summary.append(["Profit target hits", wins])
        summary.append(["Time-based square-offs", time_exits])
        summary["B3"].number_format = u'₹#,##0.00'
        wb.save(self.path)


class Controller:
    def __init__(self, config: Config, instruments: List[Instrument], mdp: MarketDataProvider):
        self.config = config
        self.instruments = [i for i in instruments if i.exchange in config.exchanges]
        self.mdp = mdp
        self.risk = RiskManager(config)
        self.strategy = Strategy52WBreakout(mdp)
        self.broker = PaperBroker(config, mdp, self.risk)

    def tick(self, now: datetime) -> List[str]:
        if self.broker.must_square_off(now):
            self.broker.square_off_all(now)
            events = list(self.broker.events)
            self.broker.events.clear()
            return events
        if self.broker.can_enter_now(now):
            for ins in self.instruments:
                sig = self.strategy.signal(ins.symbol, ins.exchange)
                if sig is not None and (ins.symbol, ins.exchange) not in self.broker.positions:
                    self.broker.place_entry(ins.symbol, ins.exchange, sig, now)
        self.broker.try_exit_targets(now)
        events = list(self.broker.events)
        self.broker.events.clear()
        return events

    def logs(self) -> List[OrderLogEntry]:
        return self.broker.logs


def run_simulation() -> str:
    config = Config(
        capital=100000.0,
        allocation_pct=0.2,
        start_time=time(9, 15),
        stop_entry_time=time(14, 30),
        square_off_time=time(15, 0),
        exchanges=["NSE", "BSE"],
        mode="mock",
    )
    mdp = MarketDataProvider(mode=config.mode)
    instruments = [
        Instrument(symbol="RELIANCE", exchange="NSE"),
        Instrument(symbol="TCS", exchange="NSE"),
        Instrument(symbol="INFY", exchange="NSE"),
    ]
    mdp.set_mock("RELIANCE", "NSE", 2900.0, 2950.0, 2200.0)
    mdp.set_mock("TCS", "NSE", 3800.0, 3750.0, 3000.0)
    mdp.set_mock("INFY", "NSE", 1550.0, 1600.0, 1200.0)
    controller = Controller(config, instruments, mdp)
    now = datetime.combine(datetime.today().date(), time(9, 20))
    mdp.update_mock_price("TCS", "NSE", 3760.0)
    controller.tick(now)
    mdp.update_mock_price("TCS", "NSE", 3760.0)
    now = datetime.combine(datetime.today().date(), time(9, 21))
    mdp.update_mock_price("TCS", "NSE", 3760.0)
    controller.tick(now)
    now = datetime.combine(datetime.today().date(), time(9, 25))
    mdp.update_mock_price("TCS", "NSE", 3787.6)
    controller.tick(now)
    now = datetime.combine(datetime.today().date(), time(15, 0))
    controller.tick(now)
    out_path = os.path.join(os.getcwd(), "transaction_log.xlsx")
    ExcelLogger(out_path).write(controller.logs())
    return out_path


def load_env(path: str = ".env") -> Dict[str, str]:
    env: Dict[str, str] = {}
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                env[k] = v
                if k not in os.environ:
                    os.environ[k] = v
    return env


def run_polling_mvp(iterations: int = 3, poll_interval_seconds: Optional[int] = None) -> Tuple[List[str], str]:
    load_env(os.path.join(os.getcwd(), ".env"))
    interval = poll_interval_seconds if poll_interval_seconds is not None else int(os.environ.get("POLL_INTERVAL_SECONDS", "10"))
    capital_env = float(os.environ.get("CAPITAL", "50000"))
    config = Config(
        capital=capital_env,
        allocation_pct=0.2,
        start_time=time(9, 15),
        stop_entry_time=time(14, 30),
        square_off_time=time(15, 0),
        exchanges=["NSE", "BSE"],
        mode=os.environ.get("MODE", "mock"),
    )
    mdp = MarketDataProvider(mode=config.mode)
    instruments = [Instrument(symbol="TCS", exchange="NSE")]
    if config.mode == "mock":
        mdp.set_mock("TCS", "NSE", 3740.0, 3750.0, 3000.0)
    controller = Controller(config, instruments, mdp)
    alerts: List[str] = []
    base_date = datetime.today().date()
    for i in range(iterations):
        now = datetime.combine(base_date, time(9, 20))
        if config.mode == "mock":
            if i == 0:
                mdp.update_mock_price("TCS", "NSE", 3755.0)
            elif i == 1:
                mdp.update_mock_price("TCS", "NSE", 3793.0)
            else:
                mdp.update_mock_price("TCS", "NSE", mdp.get_ltp("TCS", "NSE"))
        else:
            ltp = mdp.get_ltp("TCS", "NSE")
            prev = mdp.get_last_price("TCS", "NSE")
            mdp.update_mock_price("TCS", "NSE", ltp)
        events = controller.tick(now)
        alerts.extend(events)
        time_module.sleep(interval)
    now = datetime.combine(base_date, time(15, 0))
    events = controller.tick(now)
    alerts.extend(events)
    out_path = os.path.join(os.getcwd(), "transaction_log.xlsx")
    ExcelLogger(out_path).write(controller.logs())
    return alerts, out_path


if __name__ == "__main__":
    load_env(os.path.join(os.getcwd(), ".env"))
    alerts, path = run_polling_mvp(iterations=3, poll_interval_seconds=10)
    for a in alerts:
        print(a)
    print(path)

